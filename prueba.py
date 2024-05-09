from flask import Flask, render_template, request, send_file
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook
import io
import os

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def check_dni(dni):
    chrome_options = Options()
    chrome_options.add_argument("--headless")
    chrome_options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/112.0.0.0 Safari/537.36")
    chrome_options.add_argument("--disable-blink-features=AutomationControlled")
    chrome_options.add_argument("--disable-features=SameSiteByDefaultCookies,CookiesWithoutSameSiteMustBeSecure")

    chrome_options.page_load_strategy = 'eager'

    driver = webdriver.Chrome(options=chrome_options)

    try:
        # Navegar al sitio web
        print(f"Accediendo a la URL para consultar el DNI: {dni}")
        driver.get("https://e-consultaruc.sunat.gob.pe/cl-ti-itmrconsruc/FrameCriterioBusquedaWeb.jsp")
        time.sleep(5)

        # Hacer clic en "Por Documento"
        print("Haciendo clic en el botón 'Por Documento'.")
        btn_por_documento = driver.find_element(By.XPATH, '//*[@id="btnPorDocumento"]')
        btn_por_documento.click()
        time.sleep(2)

        # Ingresar el DNI
        print(f"Ingresando el DNI: {dni}.")
        txt_numero_documento = driver.find_element(By.XPATH, '//*[@id="txtNumeroDocumento"]')
        txt_numero_documento.clear()
        txt_numero_documento.send_keys(dni)

        # Hacer clic en el botón "Aceptar"
        print("Haciendo clic en el botón 'Aceptar'.")
        consultar_button = driver.find_element(By.XPATH, '//*[@id="btnAceptar"]')
        consultar_button.click()

        # Esperar la aparición de los resultados
        time.sleep(10)
        print("Extrayendo información de los resultados.")
        resultado_element = driver.find_element(By.XPATH, '/html/body/div/div[2]/div/div[3]')
        resultado_texto = resultado_element.text
        print(f"Texto de resultados obtenido:\n{resultado_texto}")

        # Analizar el texto para obtener el RUC, nombre y estado
        resultado_lineas = resultado_texto.splitlines()
        ruc_valor = resultado_lineas[1].split(':')[1].strip()
        nombre = resultado_lineas[2].strip()
        estado = resultado_lineas[4].split(':')[1].strip()

        print(f"RUC: {ruc_valor}, Nombre: {nombre}, Estado: {estado}")

    except Exception as e:
        print(f"Error durante la consulta: {e}")
        ruc_valor = "No hay resultado"
        nombre = "No hay resultado"
        estado = "No hay resultado"

    finally:
        driver.quit()
        print("Cerrando el WebDriver.")

    return ruc_valor, nombre, estado

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/result', methods=['POST'])
def result():
    numeros_dni = request.form.get('numeros_ruc').split()
    resultados = []

    for dni in numeros_dni:
        ruc_valor, nombre, estado = check_dni(dni.strip())
        resultados.append((dni.strip(), nombre, ruc_valor, estado))

    wb = Workbook()
    ws = wb.active

    ws['A1'] = 'DNI'
    ws['B1'] = 'Nombre'
    ws['C1'] = 'RUC'
    ws['D1'] = 'Estado'

    for i, (dni, nombre, ruc_valor, estado) in enumerate(resultados, start=2):
        ws.cell(row=i, column=1, value=dni)
        ws.cell(row=i, column=2, value=nombre)
        ws.cell(row=i, column=3, value=ruc_valor)
        ws.cell(row=i, column=4, value=estado)

    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)

    timestamp = time.strftime("%Y%m%d-%H%M%S")
    excel_filename = f"resultados_{timestamp}.xlsx"
    excel_path = os.path.join(UPLOAD_FOLDER, excel_filename)

    with open(excel_path, 'wb') as excel_file:
        excel_file.write(excel_io.getvalue())

    return render_template('result.html', resultados=resultados, excel_path=excel_path)

@app.route('/download_excel')
def download_excel():
    excel_path = request.args.get('excel_path', default='', type=str)

    return send_file(
        excel_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='resultados.xlsx'
    )

if __name__ == '__main__':
    app.run(debug=True)
