from flask import Flask, render_template, request, send_file
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import time
from openpyxl import Workbook
import io
import os
import time

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

def check_ruc(ruc):
    # Configurar opciones de Chrome para modo headless
    chrome_options = Options()
    chrome_options.add_argument("--headless")

    # Inicializar el driver de Selenium
    driver = webdriver.Chrome(options=chrome_options)

    try:
        # Navegar a la URL
        driver.get("https://empresasruc.com/consulta-ruc")

        # Esperar para que la página cargue completamente
        time.sleep(1)

        # Hacer clic en el radio button "RUC"
        radio_button = driver.find_element(By.XPATH, '//*[@id="radio3"]')
        radio_button.click()
        time.sleep(1)

        # Ingresar el RUC en el campo correspondiente
        num_documento_input = driver.find_element(By.XPATH, '//*[@id="content"]/div/form/div/div/div/div[1]/div/div[2]/div/div/div[3]/input')
        num_documento_input.clear()
        num_documento_input.send_keys(ruc)

        # Hacer clic en el botón 'Consultar'
        consultar_button = driver.find_element(By.XPATH, '//*[@id="content"]/div/form/div/div/div/div[1]/div/div[2]/div/div/div[3]/button')
        consultar_button.click()

        # Esperar que los resultados aparezcan
        time.sleep(6)

        # Extraer la información
        resultado_element = driver.find_element(By.XPATH, '//*[@id="modal-result-consulta-ruc"]/div/div')
        resultado_texto = resultado_element.text.split('\n')

        # Procesar la información obtenida
        informacion_relevante = []
        nombre = None
        ruc_valor = None
        estado = None

        for i, linea in enumerate(resultado_texto):
            if 'RUC:' in linea:
                ruc_valor = linea.split(':')[1].strip()
            if 'ACTIVO' in linea or 'HABIDO' in linea:
                estado = linea
            if i > 0 and 'RUC:' in resultado_texto[i - 1]:
                nombre = resultado_texto[i - 2]

    except Exception as e:
        print(f"Error durante la consulta: {e}")

    finally:
        # Asegurarse de cerrar el driver
        driver.quit()

    # Retornar la información obtenida
    if not nombre:
        nombre = "No hay resultados"
    if not ruc_valor:
        ruc_valor = "No hay resultados"
    if not estado:
        estado = "No hay resultados"

    return nombre, ruc_valor, estado, '\n'.join(informacion_relevante)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/result', methods=['POST'])
def result():
    # Obtener los números RUC ingresados por el usuario
    numeros_ruc = request.form.get('numeros_ruc')
    numeros_ruc = numeros_ruc.split()

    resultados = []

    for ruc in numeros_ruc:
        # Llamar a la función `check_ruc` para obtener los datos
        nombre, ruc_valor, estado, resultado = check_ruc(ruc.strip())

        # Reemplazar valores "None" por "No hay resultados"
        if not nombre:
            nombre = "No hay resultados"
        if not ruc_valor:
            ruc_valor = "No hay resultados"
        if not estado:
            estado = "No hay resultados"

        # Añadir los resultados a la lista de resultados
        resultados.append((ruc.strip(), nombre, ruc_valor, estado))

    # Crear un nuevo libro de Excel
    wb = Workbook()
    ws = wb.active

    # Agregar encabezado
    ws['A1'] = 'DNI'
    ws['B1'] = 'Nombre'
    ws['C1'] = 'RUC'
    ws['D1'] = 'Estado'

    # Escribir los resultados en el libro de Excel
    for i, (dni, nombre, ruc_valor, estado) in enumerate(resultados, start=2):
        ws.cell(row=i, column=1, value=dni)
        ws.cell(row=i, column=2, value=nombre)
        ws.cell(row=i, column=3, value=ruc_valor)
        ws.cell(row=i, column=4, value=estado)

    # Crear el archivo en memoria
    excel_io = io.BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)

    # Generar un nombre único usando el tiempo actual
    timestamp = time.strftime("%Y%m%d-%H%M%S")
    excel_filename = f"resultados_{timestamp}.xlsx"

    # Ruta completa del archivo Excel
    excel_path = os.path.join(UPLOAD_FOLDER, excel_filename)

    # Guardar el libro de Excel en la ruta especificada
    with open(excel_path, 'wb') as excel_file:
        excel_file.write(excel_io.getvalue())

    # Renderizar la plantilla result.html con los resultados y el enlace de descarga del archivo Excel
    return render_template('result.html', resultados=resultados, excel_path=excel_path)

@app.route('/download_excel')
def download_excel():
    excel_path = request.args.get('excel_path', default='', type=str)

    return send_file(
        excel_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='resultados.xlsx'
    )

if __name__ == '__main__':
    app.run(debug=True)
